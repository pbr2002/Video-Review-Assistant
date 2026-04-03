#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
人工辅助视频审核 GUI 工具（纯 tk 兼容版）
兼容：macOS / Python 3.9.6

本版优化：
1. 选择 Excel 文件后弹窗提示
2. 当前行号更明显（顶部 banner + 窗口标题）
3. 加载新记录时先刷新 GUI，再自动打开链接
4. 选择“不通过原因”后，直接自动判定为“审核不通过”
5. 移除“审核不通过”按钮
6. 进入下一条时自动清空原因，避免残留
"""

import re
import sys
import logging
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional, Set, Tuple
from urllib.parse import urlparse

import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

#这些是全局配置，规定默认工作表、Excel 关键列名、审核结果枚举值和不通过原因白名单。
DEFAULT_SHEET_NAME = "Sheet1"

COL_VIDEO_ID = "视频ID"
COL_VIDEO_URL = "视频链接"
COL_PLATFORM = "Unnamed: 2"
COL_REVIEW_RESULT = "*审核结果（必填；仅可选择审核通过/审核不通过）"
COL_REASON = "原因（下拉选择，填选项以外会导致上传失败；不通过原因必选！）"

REVIEW_PASS = "审核通过"
REVIEW_FAIL = "审核不通过"

REASON_WHITELIST = [
    "社区注水内容",
    "内容制作粗糙",
    "话题配置不符合活动要求",
    "投稿作品可能存在删除后投稿/作品重复投稿/文件或者链接异常等情况",
    "标题文案无效",
    "专题活动无关",
    "作品重复投稿",
    "作品曝光数据异常",
    "违规传播涉密/著作权内容",
    "违反社区发布规则",
    "不良游戏体验",
    "敏感游戏行为",
    "消极不良引导",
    "游戏相关性低",
    "有效内容不足",
    "作品时长不足，不符合活动要求",
    "图片数量不足，不符合活动要求",
    "其他",
    "复检不通过（慎用）",
]
REASON_SET = set(REASON_WHITELIST)  # type: Set[str]

DEFAULT_INVALID_LINK_REASON = "投稿作品可能存在删除后投稿/作品重复投稿/文件或者链接异常等情况"

logger = logging.getLogger("review_excel_gui")


@dataclass
#这是“当前这一行审核数据”的打包对象，里面放了行号、视频 ID、链接、平台、当前审核结果和原因。
class RowData:
    row_idx: int
    video_id: str
    video_url: str
    platform: str
    current_review_result: str
    current_reason: str

#配置日志输出，后面保存成功、跳过当前行、加载失败这些都会打印到控制台。
def setup_logging() -> None:
    logger.setLevel(logging.INFO)
    logger.handlers = []

    handler = logging.StreamHandler(sys.stdout)
    formatter = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)

#把表头文本“清洗”一下：去空格、去换行、统一格式，方便做列名匹配。
def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("\u3000", " ")
    text = text.replace("\n", "")
    text = text.replace("\r", "")
    text = re.sub(r"\s+", "", text)
    return text

#安全地把单元格内容转成字符串；如果是空值，就返回空字符串。
def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()

#建立“别名词典”。比如“视频链接”“投稿链接”“作品链接”都映射成同一个标准字段。
def build_header_alias_map() -> Dict[str, str]:
    aliases = {
        COL_VIDEO_ID: ["视频ID", "视频id"],
        COL_VIDEO_URL: ["视频链接", "链接", "作品链接", "投稿链接"],
        COL_PLATFORM: ["Unnamed: 2", "Unnamed:2", "unnamed:2", "平台"],
        COL_REVIEW_RESULT: [
            "*审核结果（必填；仅可选择审核通过/审核不通过）",
            "审核结果",
            "*审核结果",
        ],
        COL_REASON: [
            "原因（下拉选择，填选项以外会导致上传失败；不通过原因必选！）",
            "原因",
        ],
    }

    alias_map = {}  # type: Dict[str, str]
    for canonical, names in aliases.items():
        for name in names:
            alias_map[normalize_text(name)] = canonical
    return alias_map


HEADER_ALIAS_MAP = build_header_alias_map()

#把某个表头值拿来判断，它属于哪个标准列。
def match_header_name(raw_header: Any) -> Optional[str]:
    norm = normalize_text(raw_header)
    if not norm:
        return None

    if norm in HEADER_ALIAS_MAP:
        return HEADER_ALIAS_MAP[norm]

    if "视频链接" in norm:
        return COL_VIDEO_URL
    if "审核结果" in norm:
        return COL_REVIEW_RESULT
    if norm.startswith("原因"):
        return COL_REASON
    if norm in {"unnamed:2", "平台"}:
        return COL_PLATFORM
    if norm == "视频id":
        return COL_VIDEO_ID
    return None

#从 Excel 单元格里提取链接。既支持普通超链接，也支持 =HYPERLINK(...) 公式。
def extract_url_from_cell(cell: Any) -> str:
    if cell is None:
        return ""

    try:
        if cell.hyperlink and cell.hyperlink.target:
            return str(cell.hyperlink.target).strip()
    except Exception:
        pass

    value = cell.value
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    match = re.match(r'^\s*=HYPERLINK\(\s*"([^"]+)"\s*,', text, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()

    return text

#判断一个字符串看起来像不像合法的 http/https 链接。
def is_probably_url(url: str) -> bool:
    if not url:
        return False
    try:
        parsed = urlparse(url.strip())
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)
    except Exception:
        return False

#打开 Excel 文件和指定工作表；如果文件不存在、被占用、工作表不存在，就报错
def load_workbook_and_sheet(excel_path: Path, sheet_name: str) -> Tuple[Workbook, Worksheet]:
    if not excel_path.exists():
        raise FileNotFoundError("Excel 文件不存在：{0}".format(excel_path))

    try:
        wb = load_workbook(filename=str(excel_path), read_only=False, data_only=False)
    except PermissionError as exc:
        raise PermissionError(
            "无法读取文件，可能没有权限，或文件正在被其他程序占用：{0}".format(excel_path)
        ) from exc
    except Exception as exc:
        raise RuntimeError("读取 Excel 失败：{0}".format(exc)) from exc

    if sheet_name not in wb.sheetnames:
        raise ValueError("未找到工作表 {0}，现有工作表：{1}".format(sheet_name, wb.sheetnames))

    return wb, wb[sheet_name]

#按列映射把某一行的内容读出来，封装成 RowData。
def locate_columns(ws: Worksheet, header_search_rows: int = 20) -> Tuple[int, Dict[str, int]]:
    required = {COL_VIDEO_URL, COL_REVIEW_RESULT, COL_REASON}
    best_row_idx = None  # type: Optional[int]
    best_map = {}  # type: Dict[str, int]

    search_end = min(header_search_rows, ws.max_row)

    for row_idx in range(1, search_end + 1):
        current_map = {}  # type: Dict[str, int]

        for col_idx in range(1, ws.max_column + 1):
            raw_header = ws.cell(row=row_idx, column=col_idx).value
            canonical = match_header_name(raw_header)
            if canonical and canonical not in current_map:
                current_map[canonical] = col_idx

        if len(current_map) > len(best_map):
            best_map = current_map
            best_row_idx = row_idx

        if required.issubset(set(current_map.keys())):
            best_map = current_map
            best_row_idx = row_idx
            break

    if best_row_idx is None or not required.issubset(set(best_map.keys())):
        raise ValueError(
            "未能识别必要列。必要列包括：{0}；当前识别到：{1}".format(sorted(required), best_map)
        )

    if COL_VIDEO_ID not in best_map:
        logger.warning("未识别到“视频ID”列，后续将显示为空。")
    if COL_PLATFORM not in best_map:
        logger.warning("未识别到“平台”列，后续将显示为空。")

    return best_row_idx, best_map

#把审核结果和原因写回 Excel 当前行。
def get_row_data(ws: Worksheet, row_idx: int, column_map: Dict[str, int]) -> RowData:
    video_id = ""
    if COL_VIDEO_ID in column_map:
        video_id = safe_str(ws.cell(row=row_idx, column=column_map[COL_VIDEO_ID]).value)

    platform = ""
    if COL_PLATFORM in column_map:
        platform = safe_str(ws.cell(row=row_idx, column=column_map[COL_PLATFORM]).value)

    video_url = extract_url_from_cell(ws.cell(row=row_idx, column=column_map[COL_VIDEO_URL]))
    current_review_result = safe_str(ws.cell(row=row_idx, column=column_map[COL_REVIEW_RESULT]).value)
    current_reason = safe_str(ws.cell(row=row_idx, column=column_map[COL_REASON]).value)

    return RowData(
        row_idx=row_idx,
        video_id=video_id,
        video_url=video_url,
        platform=platform,
        current_review_result=current_review_result,
        current_reason=current_reason,
    )

#从开始位置往后找第一条“还没审核”的记录。
def get_next_unreviewed_row(
    ws: Worksheet,
    header_row_idx: int,
    column_map: Dict[str, int],
    start_row: Optional[int],
    skip_reviewed: bool,
) -> Optional[int]:
    review_col = column_map[COL_REVIEW_RESULT]
    begin_row = max(header_row_idx + 1, start_row or (header_row_idx + 1))

    for row_idx in range(begin_row, ws.max_row + 1):
        current_review = safe_str(ws.cell(row=row_idx, column=review_col).value)
        if skip_reviewed:
            if not current_review:
                return row_idx
        else:
            return row_idx

    return None

#从当前行往后找下一条要处理的记录。
def get_next_row_after_current(
    ws: Worksheet,
    column_map: Dict[str, int],
    current_row_idx: int,
    skip_reviewed: bool,
) -> Optional[int]:
    review_col = column_map[COL_REVIEW_RESULT]

    for row_idx in range(current_row_idx + 1, ws.max_row + 1):
        current_review = safe_str(ws.cell(row=row_idx, column=review_col).value)
        if skip_reviewed:
            if not current_review:
                return row_idx
        else:
            return row_idx

    return None

#检查审核结果和原因是否合法：
#通过时，原因必须为空
#不通过时，原因必须在白名单里
def validate_review_result(review_result: str, reason: str) -> Tuple[str, str]:
    if review_result not in {REVIEW_PASS, REVIEW_FAIL}:
        raise ValueError("非法审核结果：{0}".format(review_result))

    if review_result == REVIEW_PASS:
        if reason not in {"", None}:
            raise ValueError("审核通过时原因必须为空，当前为：{0}".format(reason))
        return REVIEW_PASS, ""

    if reason not in REASON_SET:
        raise ValueError("审核不通过时原因必须属于白名单，当前为：{0}".format(reason))

    return REVIEW_FAIL, reason


def write_result_to_row(
    ws: Worksheet,
    row_idx: int,
    column_map: Dict[str, int],
    review_result: str,
    reason: str,
) -> None:
    review_result, reason = validate_review_result(review_result, reason)

    review_col = column_map[COL_REVIEW_RESULT]
    reason_col = column_map[COL_REASON]

    ws.cell(row=row_idx, column=review_col).value = review_result
    if review_result == REVIEW_PASS:
        ws.cell(row=row_idx, column=reason_col).value = None
    else:
        ws.cell(row=row_idx, column=reason_col).value = reason

#真正执行保存动作。如果 Excel/WPS 正占着文件，就提示保存失败。
def save_workbook(wb: Workbook, excel_path: Path) -> None:
    try:
        wb.save(str(excel_path))
    except PermissionError as exc:
        raise PermissionError(
            "保存失败：文件可能正在被 Excel/WPS/Numbers 等程序占用，或当前没有写权限。\n文件：{0}".format(excel_path)
        ) from exc
    except Exception as exc:
        raise RuntimeError("保存失败：{0}".format(exc)) from exc

#尝试用默认浏览器打开视频链接；如果链接空了或格式不对，就返回失败信息。
def maybe_open_url(video_url: str) -> Tuple[bool, str]:
    if not video_url:
        return False, "当前链接为空，不打开浏览器，请手动处理。"

    if not is_probably_url(video_url):
        return False, "当前链接格式明显不合法，不打开浏览器，请手动处理。"

    try:
        opened = webbrowser.open(video_url, new=2)
        if opened:
            return True, "已尝试在默认浏览器中打开当前视频链接。"
        return False, "未能自动打开浏览器，请手动复制链接查看。"
    except Exception as exc:
        return False, "自动打开浏览器失败：{0}".format(exc)

#这是全程序最重要的类。你可以把它理解成“界面 + 主流程控制器”。它既保存当前状态，也响应按钮和下拉框事件。
class ReviewApp:
    #初始化窗口、各种状态变量，以及当前工作簿/当前行这些核心对象。
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("视频审核助手（GUI 预览版）")
        self.root.geometry("1100x720")
        self.root.minsize(980, 620)

        self.workbook = None            # type: Optional[Workbook]
        self.worksheet = None           # type: Optional[Worksheet]
        self.excel_path = None          # type: Optional[Path]
        self.column_map = None          # type: Optional[Dict[str, int]]
        self.header_row_idx = None      # type: Optional[int]
        self.current_row = None         # type: Optional[RowData]

        self.file_path_var = tk.StringVar()
        self.sheet_name_var = tk.StringVar(value=DEFAULT_SHEET_NAME)
        self.start_row_var = tk.StringVar()
        self.skip_reviewed_var = tk.BooleanVar(value=True)
        self.auto_open_next_var = tk.BooleanVar(value=True)

        self.selected_file_name_var = tk.StringVar(value="未选择文件")
        self.current_row_banner_var = tk.StringVar(value="当前行：-")
        self.row_idx_var = tk.StringVar(value="-")
        self.video_id_var = tk.StringVar(value="-")
        self.platform_var = tk.StringVar(value="-")
        self.video_url_var = tk.StringVar(value="-")
        self.current_review_var = tk.StringVar(value="-")
        self.current_reason_var = tk.StringVar(value="-")
        self.progress_var = tk.StringVar(value="未加载文件")
        self.status_var = tk.StringVar(value="请先选择 Excel 文件")

        self.reason_var = tk.StringVar(value="")
        self._is_loading_reason = False

        self.BG = "#f5f5f5"
        self.FG = "#111111"
        self.PANEL_BG = "#ffffff"

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def _make_entry(self, parent, textvariable=None, width=None):
        entry = tk.Entry(
            parent,
            textvariable=textvariable,
            width=width,
            bg="white",
            fg="black",
            insertbackground="black",
            relief="solid",
            bd=1,
        )
        return entry

    def _make_button(self, parent, text, command, width=None):
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg="#ffffff",
            fg="#111111",
            relief="raised",
            bd=1,
            padx=8,
            pady=4,
            width=width,
        )
        return btn

    def _make_value_label(self, parent, textvariable, wraplength=None, justify="left", font=None):
        lbl = tk.Label(
            parent,
            textvariable=textvariable,
            bg=self.PANEL_BG,
            fg=self.FG,
            anchor="w",
            justify=justify,
            wraplength=wraplength,
            font=font,
        )
        return lbl
    #把整个界面搭出来：文件选择区、参数区、当前记录区、操作区、状态栏。
    def _build_ui(self) -> None:
        self.root.configure(bg=self.BG)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(3, weight=1)

        top = tk.Frame(self.root, bg=self.BG, padx=12, pady=12)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        tk.Label(top, text="Excel 文件：", bg=self.BG, fg=self.FG).grid(
            row=0, column=0, sticky="w", padx=(0, 8)
        )
        self._make_entry(top, textvariable=self.file_path_var).grid(
            row=0, column=1, sticky="ew"
        )
        self._make_button(top, "选择文件", self.choose_file).grid(
            row=0, column=2, padx=(8, 0)
        )
        self._make_button(top, "加载并开始", self.start_review).grid(
            row=0, column=3, padx=(8, 0)
        )

        tk.Label(
            top,
            textvariable=self.selected_file_name_var,
            bg=self.BG,
            fg="#555555",
            anchor="w",
        ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(8, 0))

        option = tk.Frame(self.root, bg=self.BG, padx=12, pady=8)
        option.grid(row=1, column=0, sticky="ew")
        option.grid_columnconfigure(7, weight=1)

        tk.Label(option, text="工作表：", bg=self.BG, fg=self.FG).grid(
            row=0, column=0, sticky="w"
        )
        self._make_entry(option, textvariable=self.sheet_name_var, width=12).grid(
            row=0, column=1, sticky="w", padx=(4, 12)
        )

        tk.Label(option, text="起始行：", bg=self.BG, fg=self.FG).grid(
            row=0, column=2, sticky="w"
        )
        self._make_entry(option, textvariable=self.start_row_var, width=10).grid(
            row=0, column=3, sticky="w", padx=(4, 12)
        )

        tk.Checkbutton(
            option,
            text="跳过已审核行",
            variable=self.skip_reviewed_var,
            bg=self.BG,
            fg=self.FG,
            activebackground=self.BG,
            activeforeground=self.FG,
            selectcolor="white",
        ).grid(row=0, column=4, sticky="w", padx=(0, 12))

        tk.Checkbutton(
            option,
            text="进入新记录时自动打开链接",
            variable=self.auto_open_next_var,
            bg=self.BG,
            fg=self.FG,
            activebackground=self.BG,
            activeforeground=self.FG,
            selectcolor="white",
        ).grid(row=0, column=5, sticky="w", padx=(0, 12))

        banner = tk.Frame(self.root, bg="#ffffff", bd=1, relief="solid", padx=12, pady=10)
        banner.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 8))
        banner.columnconfigure(0, weight=1)

        tk.Label(
            banner,
            textvariable=self.current_row_banner_var,
            bg="#ffffff",
            fg="#111111",
            anchor="w",
            font=("Arial", 14, "bold"),
        ).grid(row=0, column=0, sticky="w")

        info = tk.LabelFrame(
            self.root,
            text="当前记录",
            bg=self.PANEL_BG,
            fg=self.FG,
            padx=12,
            pady=12,
            bd=1,
            relief="solid",
        )
        info.grid(row=3, column=0, sticky="nsew", padx=12, pady=(0, 8))
        info.columnconfigure(1, weight=1)

        tk.Label(info, text="当前行号：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=0, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.row_idx_var).grid(
            row=0, column=1, sticky="nw", pady=4
        )

        tk.Label(info, text="视频ID：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=1, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.video_id_var).grid(
            row=1, column=1, sticky="nw", pady=4
        )

        tk.Label(info, text="平台：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=2, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.platform_var).grid(
            row=2, column=1, sticky="nw", pady=4
        )

        tk.Label(info, text="视频链接：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=3, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.video_url_var, wraplength=820, justify="left").grid(
            row=3, column=1, sticky="nw", pady=4
        )

        tk.Label(info, text="当前已有审核结果：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=4, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.current_review_var).grid(
            row=4, column=1, sticky="nw", pady=4
        )

        tk.Label(info, text="当前已有原因：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=5, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.current_reason_var).grid(
            row=5, column=1, sticky="nw", pady=4
        )

        tk.Label(info, text="进度：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=6, column=0, sticky="nw", pady=4
        )
        self._make_value_label(info, self.progress_var).grid(
            row=6, column=1, sticky="nw", pady=4
        )

        action = tk.LabelFrame(
            self.root,
            text="操作区",
            bg=self.PANEL_BG,
            fg=self.FG,
            padx=12,
            pady=12,
            bd=1,
            relief="solid",
        )
        action.grid(row=4, column=0, sticky="ew", padx=12, pady=(0, 8))
        action.columnconfigure(1, weight=1)

        self._make_button(action, "打开当前视频", self.on_open_current).grid(
            row=0, column=0, padx=(0, 8), pady=(0, 8), sticky="w"
        )
        self._make_button(action, "审核通过", self.on_pass).grid(
            row=0, column=1, padx=(0, 8), pady=(0, 8), sticky="w"
        )
        self._make_button(action, "跳过当前行", self.on_skip).grid(
            row=0, column=2, padx=(0, 8), pady=(0, 8), sticky="w"
        )
        self._make_button(action, "保存并退出", self.on_save_and_quit).grid(
            row=0, column=3, pady=(0, 8), sticky="w"
        )

        tk.Label(action, text="不通过原因：", bg=self.PANEL_BG, fg=self.FG).grid(
            row=1, column=0, sticky="w"
        )

        reason_choices = [""] + REASON_WHITELIST
        self.reason_menu = tk.OptionMenu(action, self.reason_var, *reason_choices)
        self.reason_menu.config(
            bg="white",
            fg="black",
            activebackground="#eaeaea",
            activeforeground="black",
            relief="solid",
            bd=1,
            highlightthickness=0,
            anchor="w",
            width=72,
        )
        self.reason_menu["menu"].config(
            bg="white",
            fg="black",
            activebackground="#eaeaea",
            activeforeground="black",
        )
        self.reason_menu.grid(row=1, column=1, columnspan=3, sticky="ew")

        tk.Label(
            action,
            text="提示：选择一个不通过原因后，将自动判定为“审核不通过”并跳转下一条。",
            bg=self.PANEL_BG,
            fg="#666666",
            anchor="w",
        ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(8, 0))

        self.reason_var.trace_add("write", self.on_reason_selected)

        bottom = tk.Frame(self.root, bg=self.BG, padx=12, pady=12)
        bottom.grid(row=5, column=0, sticky="ew")
        bottom.columnconfigure(0, weight=1)

        tk.Label(
            bottom,
            textvariable=self.status_var,
            bg=self.BG,
            fg="#333333",
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
    #点“选择文件”后弹出文件选择框，并把选中的路径显示到界面上。
    def choose_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("所有文件", "*.*"),
            ],
        )

        if not file_path:
            self.status_var.set("未选择文件。")
            return

        self.file_path_var.set(file_path)
        self.selected_file_name_var.set("已选择文件：{0}".format(Path(file_path).name))
        self.status_var.set("已选择 Excel 文件，请点击“加载并开始”。")
        self.root.update_idletasks()

        messagebox.showinfo(
            "已选择文件",
            "已选择：\n{0}\n\n请继续点击“加载并开始”。".format(file_path)
        )
    #把用户填的“起始行”解析成数字，并校验是否合法。
    def parse_start_row(self) -> Optional[int]:
        text = self.start_row_var.get().strip()
        if not text:
            return None
        if not text.isdigit():
            raise ValueError("起始行必须为空或正整数。")
        value = int(text)
        if value <= 0:
            raise ValueError("起始行必须为空或正整数。")
        return value
    '''这是“加载并开始”的总入口，负责：
        检查文件有没有选
        解析工作表和起始行
        确认会覆盖原文件
        打开 Excel
        定位关键列
        找第一条待审核记录
        调用 load_row() 加载它。'''
    def start_review(self) -> None:
        try:
            path_text = self.file_path_var.get().strip()
            if not path_text:
                messagebox.showwarning("提示", "请先选择 Excel 文件。")
                return

            excel_path = Path(path_text).expanduser().resolve()
            sheet_name = self.sheet_name_var.get().strip() or DEFAULT_SHEET_NAME
            start_row = self.parse_start_row()

            confirm = messagebox.askyesno(
                "确认覆盖",
                "该工具会直接覆盖原 Excel 文件。\n\n文件：{0}\n\n是否继续？".format(excel_path),
            )
            if not confirm:
                self.status_var.set("已取消，不会修改原文件。")
                return

            workbook, worksheet = load_workbook_and_sheet(excel_path, sheet_name)
            header_row_idx, column_map = locate_columns(worksheet, header_search_rows=20)

            self.workbook = workbook
            self.worksheet = worksheet
            self.excel_path = excel_path
            self.header_row_idx = header_row_idx
            self.column_map = column_map

            next_row = get_next_unreviewed_row(
                ws=self.worksheet,
                header_row_idx=self.header_row_idx,
                column_map=self.column_map,
                start_row=start_row,
                skip_reviewed=self.skip_reviewed_var.get(),
            )

            if next_row is None:
                self.clear_current_display()
                self.status_var.set("未找到可处理的记录。")
                messagebox.showinfo("提示", "从当前起点开始，未找到可处理的记录。")
                return

            self.load_row(next_row)
            self.status_var.set("加载成功，开始审核。")

        except Exception as exc:
            logger.exception("加载失败：%s", exc)
            messagebox.showerror("错误", str(exc))
            self.status_var.set("加载失败：{0}".format(exc))
    #把某一条记录真正显示到界面上，同时更新 banner、标题、进度；如果开启自动打开链接，还会自动打开浏览器。
    def load_row(self, row_idx: int) -> None:
        if self.worksheet is None or self.column_map is None:
            return

        row = get_row_data(self.worksheet, row_idx, self.column_map)
        self.current_row = row

        self.row_idx_var.set(str(row.row_idx))
        self.video_id_var.set(row.video_id or "空")
        self.platform_var.set(row.platform or "空")
        self.video_url_var.set(row.video_url or "空")
        self.current_review_var.set(row.current_review_result or "空")
        self.current_reason_var.set(row.current_reason or "空")

        self._is_loading_reason = True
        self.reason_var.set("")
        self._is_loading_reason = False

        max_row = self.worksheet.max_row if self.worksheet is not None else 0
        self.progress_var.set("当前 Excel 行：{0} / {1}".format(row.row_idx, max_row))
        self.current_row_banner_var.set(
            "当前行：{0}    视频ID：{1}".format(row.row_idx, row.video_id or "空")
        )
        self.status_var.set(
            "当前正在处理第 {0} 行，视频ID：{1}".format(row.row_idx, row.video_id or "空")
        )
        self.root.title("视频审核助手（GUI 预览版） - 当前第 {0} 行".format(row.row_idx))

        self.root.update_idletasks()

        if self.auto_open_next_var.get():
            self.root.after(200, self.open_current_url)
    #当没有记录可处理了，或者还没加载文件时，把界面恢复成空状态。
    def clear_current_display(self) -> None:
        self.current_row = None
        self.row_idx_var.set("-")
        self.video_id_var.set("-")
        self.platform_var.set("-")
        self.video_url_var.set("-")
        self.current_review_var.set("-")
        self.current_reason_var.set("-")
        self.progress_var.set("已完成或未加载")
        self.current_row_banner_var.set("当前行：-")
        self.root.title("视频审核助手（GUI 预览版）")

        self._is_loading_reason = True
        self.reason_var.set("")
        self._is_loading_reason = False
    #打开当前记录的视频链接；如果链接无效，会提示“建议判为哪种不通过原因”。
    def open_current_url(self) -> None:
        if self.current_row is None:
            messagebox.showinfo("提示", "当前没有可处理的记录。")
            return

        opened, msg = maybe_open_url(self.current_row.video_url)
        self.status_var.set(msg)

        if not opened:
            if not self.current_row.video_url or not is_probably_url(self.current_row.video_url):
                messagebox.showinfo(
                    "链接提示",
                    "{0}\n\n如需判为不通过，建议原因：\n{1}".format(msg, DEFAULT_INVALID_LINK_REASON),
                )
    #“打开当前视频”按钮的事件函数，本质就是调用 open_current_url()。
    def on_open_current(self) -> None:
        self.open_current_url()
    #这是全程序最值钱的一个函数。它监听“不通过原因”的变化，只要用户选了一个合法原因，就自动触发保存和跳转。
    def on_reason_selected(self, *args) -> None:
        if self._is_loading_reason:
            return

        if self.current_row is None:
            return

        reason = self.reason_var.get().strip()
        if not reason:
            return

        if reason not in REASON_SET:
            self.status_var.set("选择的原因不合法，请重新选择。")
            return

        self.status_var.set(
            "第 {0} 行已选择不通过原因：{1}，正在自动保存...".format(
                self.current_row.row_idx, reason
            )
        )
        self.root.update_idletasks()
        self.save_current_result_and_go_next(REVIEW_FAIL, reason)
    '''这是“保存并流转”的核心函数。它会：
        检查当前是否已加载文件和当前行
        校验审核结果
        写回 Excel
        保存文件
        找下一条
        加载下一条，或者提示全部完成。'''
    def save_current_result_and_go_next(self, review_result: str, reason: str) -> None:
        if self.workbook is None or self.worksheet is None or self.excel_path is None or self.column_map is None:
            messagebox.showwarning("提示", "请先加载 Excel 文件。")
            return

        if self.current_row is None:
            messagebox.showinfo("提示", "当前没有可处理的记录。")
            return

        try:
            review_result, reason = validate_review_result(review_result, reason)
            current_row_idx = self.current_row.row_idx

            write_result_to_row(
                ws=self.worksheet,
                row_idx=current_row_idx,
                column_map=self.column_map,
                review_result=review_result,
                reason=reason,
            )

            save_workbook(self.workbook, self.excel_path)
            logger.info(
                "保存成功 | row=%s | review_result=%s | reason=%s",
                current_row_idx,
                review_result,
                reason,
            )
            self.status_var.set("第 {0} 行已保存。".format(current_row_idx))

            next_row = get_next_row_after_current(
                ws=self.worksheet,
                column_map=self.column_map,
                current_row_idx=current_row_idx,
                skip_reviewed=self.skip_reviewed_var.get(),
            )

            if next_row is None:
                self.clear_current_display()
                self.status_var.set("全部处理完成。")
                messagebox.showinfo("完成", "全部可处理记录已完成。")
                return

            self.load_row(next_row)

        except Exception as exc:
            logger.exception("保存失败：%s", exc)
            messagebox.showerror("错误", str(exc))
            self.status_var.set("保存失败：{0}".format(exc))
    #点“审核通过”按钮时，调用统一保存函数，并传入“通过 + 空原因”。
    def on_pass(self) -> None:
        self.save_current_result_and_go_next(REVIEW_PASS, "")
    #跳过当前行，不写结果，直接切到下一条。
    def on_skip(self) -> None:
        if self.worksheet is None or self.column_map is None or self.current_row is None:
            messagebox.showinfo("提示", "当前没有可处理的记录。")
            return

        current_row_idx = self.current_row.row_idx
        next_row = get_next_row_after_current(
            ws=self.worksheet,
            column_map=self.column_map,
            current_row_idx=current_row_idx,
            skip_reviewed=self.skip_reviewed_var.get(),
        )

        logger.info("用户跳过当前行 | row=%s", current_row_idx)

        if next_row is None:
            self.clear_current_display()
            self.status_var.set("全部处理完成。")
            messagebox.showinfo("完成", "全部可处理记录已完成。")
            return

        self.load_row(next_row)
    #退出前先保存，再关闭窗口。
    def on_save_and_quit(self) -> None:
        try:
            if self.workbook is not None and self.excel_path is not None:
                save_workbook(self.workbook, self.excel_path)
            self.root.destroy()
        except Exception as exc:
            messagebox.showerror("错误", str(exc))
    #用户点窗口右上角关闭时，先弹确认框，再决定是否退出。
    def on_close(self) -> None:
        confirm = messagebox.askyesno("退出确认", "确定退出吗？")
        if not confirm:
            return
        self.on_save_and_quit()

#初始化日志，创建 Tkinter 根窗口，实例化 ReviewApp，最后启动 mainloop()。
def main() -> int:
    setup_logging()
    root = tk.Tk()
    app = ReviewApp(root)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())